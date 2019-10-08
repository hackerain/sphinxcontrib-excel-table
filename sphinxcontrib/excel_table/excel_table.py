import os
import re
import json
import datetime

import docutils.core
from docutils.parsers.rst import Directive
from docutils.parsers.rst import directives

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader


def _get_resource_dir(folder):
    current_path = os.path.dirname(__file__)
    resource_path = os.path.join(current_path, folder)
    return resource_path

def az_to_dec(s):
    if not s:
        return 0

    n = 0
    i = len(s) -1
    j = 1

    while True:
        if i < 0:
            break
        c = s[i].upper()
        n += (ord(c) % 32) * j
        i -= 1
        j *= 26

    return n

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime.date):
            date_string = obj.strftime('%Y-%m-%d')
            return date_string
        if isinstance(obj, datetime.datetime):
            datetime_string = obj.strftime("%Y-%m-%d %H:%M:%S")
            return datetime_string
        return json.JSONEncoder.default(self, obj)

def _dumps(data):
    return json.dumps(data, cls=DateTimeEncoder)

class ExcelTable(Directive):

    option_spec = {
        'file': directives.path,
        'sheet': directives.unchanged,
        'rows': directives.unchanged,
        'selection': directives.unchanged,
        'overflow': directives.unchanged,
        'tablewidth': directives.unchanged,
        'colwidths': directives.unchanged,
        'row_header': directives.unchanged,
        'col_header': directives.unchanged,
    }

    def run(self, icnt=[0]):
        env = self.state.document.settings.env
        document = self.state.document

        # Track the number of times the directive is called,
        # send this to jinja template
        icnt[0] += 1

        file_path = self.options.get('file')
        selection = self.options.get('selection')
        rows = self.options.get('rows')
        sheet_name = self.options.get('sheet')
        overflow = self.options.get('overflow', 'horizontal')
        tablewidth = self.options.get('tablewidth', 'undefined')
        colwidths = self.options.get('colwidths', 'undefined')
        row_header = self.options.get('row_header', 'true')
        col_header = self.options.get('col_header', 'true')

        if not file_path:
            msg = "file option is missing"
            return [document.reporter.warning(msg, line=self.lineno)]

        if selection and ':' not in selection:
            msg = "selection must contain a range seperated by :"
            return [document.reporter.warning(msg, line=self.lineno)]

        if rows and ':' not in rows:
            msg = "rows must contain a range seperated by :"
            return [document.reporter.warning(msg, line=self.lineno)]

        relfn, excel_file = env.relfn2path(file_path)
        env.note_dependency(relfn)

        if overflow != 'false':
            overflow = "\'%s\'" % overflow

        data = {
          'file_name': file_path,
          'sheet_name': sheet_name,
          'overflow': overflow,
          'tablewidth': tablewidth,
          'colwidths': colwidths,
          'row_header': row_header,
          'col_header': col_header,
          'icnt': icnt[0],
        }

        wb = load_workbook(filename=excel_file)

        if sheet_name and sheet_name not in wb.sheetnames:
            msg = "sheet %s does not exist" % sheet_name
            return [document.reporter.warning(msg, line=self.lineno)]

        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.worksheets[0]

        if selection:
            sheet_data = sheet[selection]
        elif rows:
            sheet_data = sheet[rows]
        else:
            sheet_data = sheet

        content = []
        for row in sheet_data:
            _row = []
            for cell in row:
                _row.append(cell.value)
            content.append(_row)

        data['content'] = _dumps(content)

        cell_regex = re.compile(r'([A-Z]+){1}([0-9]+){1}')
        merged_cells = []

        for m in sheet.merged_cells:
            start, end = str(m).split(":")

            start_match = cell_regex.search(start)
            start_col, start_row = start_match.groups()

            end_match = cell_regex.search(end)
            end_col, end_row = end_match.groups()

            m_row = int(start_row) - 1
            m_col = az_to_dec(start_col) - 1
            m_rowspan = int(end_row) - int(start_row) + 1
            m_colspan = az_to_dec(end_col) - az_to_dec(start_col) + 1

            merged_cells.append({
                "row": m_row,
                "col": m_col,
                "rowspan": m_rowspan,
                "colspan": m_colspan,
            })

        data['merged_cells'] = _dumps(merged_cells)

        loader = FileSystemLoader(_get_resource_dir('templates'))
        _env = Environment(loader=loader,
                           keep_trailing_newline=True,
                           trim_blocks=True,
                           lstrip_blocks=True)

        template = _env.get_template('table.html')
        html = template.render(**data)
        return [docutils.nodes.raw('', html, format='html')]
